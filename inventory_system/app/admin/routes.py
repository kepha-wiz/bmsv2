from flask import render_template, redirect, url_for, flash, request, jsonify, make_response
from flask_login import login_required, current_user
from datetime import datetime, timedelta
from sqlalchemy import func, desc, and_
from io import BytesIO
import pandas as pd
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import tempfile
import os

from app import db
from app.models import Product, StockAddition, Sale, User
from app.admin.forms import ProductForm, StockAdditionForm, ReportForm
from app.admin import bp

# Create a custom PageTemplate for footer
class FooterPageTemplate(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        
    def footer(self, doc):
        # Save the state
        self.saveState()
        
        # Footer
        footer = Paragraph("© Copyright CKS Tech | Contact: +256 755261254 | Email: kephacheps55@gmail.com", 
                          getSampleStyleSheet()['Normal'])
        footer.wrap(doc.width, doc.bottomMargin)
        footer.drawOn(self, doc.leftMargin, doc.bottomMargin)
        
        # Page number
        self.setFont('Helvetica', 9)
        self.setFillColor(colors.black)
        self.drawRightString(doc.pagesize[0] - doc.rightMargin, 
                            doc.bottomMargin, 
                            f"Page {self.getPageNumber()}")
        
        # Restore the state
        self.restoreState()

# Custom DocTemplate with footer
class DocTemplateWithFooter(SimpleDocTemplate):
    def __init__(self, *args, **kwargs):
        SimpleDocTemplate.__init__(self, *args, **kwargs)
        
    def afterPage(self):
        # Add footer to each page
        canvas = self.canv
        canvas.saveState()
        
        # Footer line
        canvas.setStrokeColorRGB(0.8, 0.8, 0.8)
        canvas.setLineWidth(1)
        canvas.line(self.leftMargin, self.bottomMargin + 0.5 * inch, 
                   self.pagesize[0] - self.rightMargin, self.bottomMargin + 0.5 * inch)
        
        # Footer text
        canvas.setFont('Helvetica', 9)
        canvas.setFillColorRGB(0.4, 0.4, 0.4)
        canvas.drawString(self.leftMargin, self.bottomMargin + 0.2 * inch, 
                          "© Copyright CKS Tech | Contact: +256 755261254 | Email: kephacheps55@gmail.com")
        
        # Page number
        canvas.drawRightString(self.pagesize[0] - self.rightMargin, 
                              self.bottomMargin + 0.2 * inch, 
                              f"Page {self.page}")
        
        canvas.restoreState()

@bp.before_request
def admin_required():
    if not current_user.is_authenticated or not current_user.is_admin():
        flash('Admin access required!', 'danger')
        return redirect(url_for('auth.login'))

@bp.route('/dashboard')
@login_required
def dashboard():
    # Calculate statistics
    total_products = Product.query.count()
    total_stock_value = db.session.query(func.sum(Product.quantity_in_stock * Product.cost_price)).scalar() or 0
    total_sales = db.session.query(func.sum(Sale.total_amount)).scalar() or 0
    low_stock_products = Product.query.filter(Product.quantity_in_stock < 10).count()
    
    # Sales summary for last 30 days
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)
    recent_sales = Sale.query.filter(Sale.timestamp >= thirty_days_ago).count()
    
    # Best selling products
    best_sellers = db.session.query(
        Product.name,
        func.sum(Sale.quantity_sold).label('total_sold')
    ).join(Sale).group_by(Product.id).order_by(desc('total_sold')).limit(5).all()
    
    # Recent sales activities
    recent_activities = Sale.query.order_by(desc(Sale.timestamp)).limit(10).all()
    
    # Low stock products
    low_stock = Product.query.filter(Product.quantity_in_stock < 10).all()
    
    # Sales data for chart (last 7 days)
    seven_days_ago = datetime.utcnow() - timedelta(days=7)
    sales_chart_data = db.session.query(
        func.date(Sale.timestamp).label('date'),
        func.sum(Sale.total_amount).label('total')
    ).filter(Sale.timestamp >= seven_days_ago).group_by(func.date(Sale.timestamp)).all()
    
    dates = [str(item.date) for item in sales_chart_data]
    amounts = [float(item.total) for item in sales_chart_data]
    
    return render_template('admin/dashboard.html',
                         total_products=total_products,
                         total_stock_value=total_stock_value,
                         total_sales=total_sales,
                         low_stock_products=low_stock_products,
                         recent_sales=recent_sales,
                         best_sellers=best_sellers,
                         recent_activities=recent_activities,
                         low_stock=low_stock,
                         dates=dates,
                         amounts=amounts)

@bp.route('/stock_movement')
@login_required
def stock_movement():
    # Get all products
    products = Product.query.all()
    
    stock_data = []
    for product in products:
        # Calculate opening stock (initial stock when product was first added)
        opening_stock = 0
        
        # Get all stock additions for this product
        stock_additions = StockAddition.query.filter_by(product_id=product.id).all()
        total_added = sum([sa.quantity_added for sa in stock_additions])
        
        # Get all sales for this product
        sales = Sale.query.filter_by(product_id=product.id).all()
        total_sold = sum([s.quantity_sold for s in sales])
        
        # Calculate current balance
        current_balance = opening_stock + total_added - total_sold
        
        stock_data.append({
            'product': product,
            'opening_stock': opening_stock,
            'total_added': total_added,
            'total_sold': total_sold,
            'current_balance': current_balance
        })
    
    return render_template('admin/stock_movement.html', stock_data=stock_data)

@bp.route('/products')
@login_required
def products():
    products = Product.query.all()
    return render_template('admin/products.html', products=products)

@bp.route('/add_product', methods=['GET', 'POST'])
@login_required
def add_product():
    form = ProductForm()
    
    if form.validate_on_submit():
        product = Product(
            name=form.name.data,
            sku=form.sku.data,
            cost_price=form.cost_price.data,
            selling_price=form.selling_price.data,
            quantity_in_stock=form.quantity_in_stock.data,
            category=form.category.data,
            description=form.description.data,
            low_stock_threshold=form.low_stock_threshold.data
        )
        db.session.add(product)
        db.session.commit()
        flash('Product added successfully!', 'success')
        return redirect(url_for('admin.products'))
    
    return render_template('admin/add_product.html', form=form)

@bp.route('/add_stock', methods=['GET', 'POST'])
@login_required
def add_stock():
    form = StockAdditionForm()
    
    if form.validate_on_submit():
        product = Product.query.get(form.product_id.data)
        if product:
            # Track if prices were updated
            price_updated = False
            old_cost_price = product.cost_price
            old_selling_price = product.selling_price
            
            # Update prices if provided
            if form.cost_price.data and form.cost_price.data != product.cost_price:
                product.cost_price = form.cost_price.data
                price_updated = True
                
            if form.selling_price.data and form.selling_price.data != product.selling_price:
                product.selling_price = form.selling_price.data
                price_updated = True
            
            # Add stock
            stock_addition = StockAddition(
                product_id=product.id,
                quantity_added=form.quantity.data,
                added_by=current_user.id,
                old_cost_price=old_cost_price,
                old_selling_price=old_selling_price,
                new_cost_price=form.cost_price.data if form.cost_price.data else None,
                new_selling_price=form.selling_price.data if form.selling_price.data else None,
                price_change_reason=form.reason.data if price_updated else None
            )
            db.session.add(stock_addition)
            
            # Update product quantity
            product.quantity_in_stock += form.quantity.data
            db.session.commit()
            
            # Create appropriate success message
            if price_updated:
                if form.cost_price.data and form.selling_price.data:
                    flash(f'Successfully added {form.quantity.data} units to {product.name} and updated both prices!', 'success')
                elif form.cost_price.data:
                    flash(f'Successfully added {form.quantity.data} units to {product.name} and updated cost price from UGX{old_cost_price:.2f} to UGX{form.cost_price.data:.2f}!', 'success')
                elif form.selling_price.data:
                    flash(f'Successfully added {form.quantity.data} units to {product.name} and updated selling price from UGX{old_selling_price:.2f} to UGX{form.selling_price.data:.2f}!', 'success')
            else:
                flash(f'Successfully added {form.quantity.data} units to {product.name}!', 'success')
                
            return redirect(url_for('admin.products'))
    
    return render_template('admin/add_stock.html', form=form)

@bp.route('/search_products')
@login_required
def search_products():
    query = request.args.get('query', '').strip()
    
    if not query:
        return redirect(url_for('admin.products'))
    
    # Search for products by name or SKU
    products = Product.query.filter(
        db.or_(
            Product.name.contains(query),
            Product.sku.contains(query)
        )
    ).all()
    
    return render_template('admin/search_results.html', 
                         products=products, 
                         query=query,
                         search_type='admin')

@bp.route('/advanced_search')
@login_required
def advanced_search():
    query = request.args.get('query', '').strip()
    category = request.args.get('category', '')
    min_stock = request.args.get('min_stock', type=int)
    max_price = request.args.get('max_price', type=float)
    min_price = request.args.get('min_price', type=float)
    in_stock_only = request.args.get('in_stock_only') == 'on'
    
    # Build the query
    filters = []
    
    if query:
        filters.append(
            db.or_(
                Product.name.contains(query),
                Product.sku.contains(query)
            )
        )
    
    if category:
        filters.append(Product.category == category)
    
    if min_stock is not None:
        filters.append(Product.quantity_in_stock >= min_stock)
    
    if max_price is not None:
        filters.append(Product.selling_price <= max_price)
    
    if min_price is not None:
        filters.append(Product.selling_price >= min_price)
    
    if in_stock_only:
        filters.append(Product.quantity_in_stock > 0)
    
    # Apply all filters
    products = Product.query.filter(db.and_(*filters)).all() if filters else Product.query.all()
    
    return render_template('admin/advanced_search_results.html', 
                         products=products, 
                         query=query,
                         category=category,
                         min_stock=min_stock,
                         max_price=max_price,
                         min_price=min_price,
                         in_stock_only=in_stock_only)

@bp.route('/stock_history/<int:product_id>')
@login_required
def stock_history(product_id):
    product = Product.query.get_or_404(product_id)
    
    # Get all stock additions for this product
    stock_additions = StockAddition.query.filter_by(product_id=product_id).order_by(StockAddition.date_added.desc()).all()
    
    return render_template('admin/stock_history.html', product=product, stock_additions=stock_additions)

@bp.route('/sales')
@login_required
def sales():
    sales = Sale.query.order_by(desc(Sale.timestamp)).all()
    return render_template('admin/sales.html', sales=sales)

@bp.route('/reports')
@login_required
def reports():
    form = ReportForm()
    return render_template('admin/reports.html', form=form)

@bp.route('/generate_report/<report_type>', methods=['POST'])
@login_required
def generate_report(report_type):
    form = ReportForm()
    if not form.validate_on_submit():
        flash('Invalid date range selected', 'danger')
        return redirect(url_for('admin.reports'))
    
    start_date = form.start_date.data
    end_date = form.end_date.data
    export_format = request.form.get('export_format', 'pdf')  # Get export format from form
    
    if report_type == 'sales':
        if export_format == 'excel':
            return generate_sales_excel(start_date, end_date)
        else:
            return generate_sales_report(start_date, end_date)
    elif report_type == 'stock':
        if export_format == 'excel':
            return generate_stock_excel(start_date, end_date)
        else:
            return generate_stock_report(start_date, end_date)
    else:
        flash('Invalid report type', 'danger')
        return redirect(url_for('admin.reports'))

def generate_sales_report(start_date, end_date):
    # Query sales data
    sales_data = db.session.query(
        Sale.id,
        Product.name.label('product_name'),
        Product.sku,
        Product.category,
        Sale.quantity_sold,
        Sale.price_per_unit,
        Sale.total_amount,
        Sale.timestamp,
        User.username.label('employee')
    ).join(Product).join(User).filter(
        and_(Sale.timestamp >= start_date, Sale.timestamp <= end_date)
    ).all()
    
    # Generate PDF
    buffer = BytesIO()
    doc = DocTemplateWithFooter(
        buffer, 
        pagesize=landscape(letter),
        leftMargin=0.75*inch,
        rightMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=1*inch
    )
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2E4057')
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#4A6572')
    )
    
    # Company header
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#546E7A')
    )
    
    # Add company header
    elements.append(Paragraph("CKS BUSINESS MANAGEMENT SYSTEM", header_style))
    elements.append(Spacer(1, 0.1*inch))
    
    # Add title
    elements.append(Paragraph("SALES REPORT", title_style))
    elements.append(Paragraph(f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Report generation info
    report_info_style = ParagraphStyle(
        'ReportInfo',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_RIGHT,
        textColor=colors.HexColor('#546E7A')
    )
    
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", report_info_style))
    elements.append(Paragraph(f"Generated by: {current_user.username}", report_info_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary section
    elements.append(Paragraph("REPORT SUMMARY", subtitle_style))
    
    # Calculate summary data
    total_sales = sum([sale.total_amount for sale in sales_data])
    total_quantity = sum([sale.quantity_sold for sale in sales_data])
    total_transactions = len(sales_data)
    
    # Create summary table
    summary_data = [
        ['Total Sales Amount', f'UGX {total_sales:,.2f}'],
        ['Total Units Sold', f'{total_quantity:,} units'],
        ['Number of Transactions', f'{total_transactions:,}'],
        ['Average Transaction Value', f'UGX {total_sales/total_transactions if total_transactions > 0 else 0:,.2f}']
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#424242')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0'))
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Sales details section
    elements.append(Paragraph("SALES DETAILS", subtitle_style))
    
    if sales_data:
        # Create sales data table
        headers = ['ID', 'Product', 'SKU', 'Category', 'Qty', 'Price/Unit', 'Total', 'Date', 'Employee']
        data = [headers]
        
        for sale in sales_data:
            data.append([
                str(sale.id),
                sale.product_name,
                sale.sku,
                sale.category,
                str(sale.quantity_sold),
                f'UGX {sale.price_per_unit:,.2f}',
                f'UGX {sale.total_amount:,.2f}',
                sale.timestamp.strftime('%Y-%m-%d %H:%M'),
                sale.employee
            ])
        
        # Create table with alternating row colors
        table = Table(data, colWidths=[0.5*inch, 1.5*inch, 0.8*inch, 1*inch, 0.5*inch, 1*inch, 1*inch, 1.2*inch, 1*inch])
        
        # Style the table
        row_colors = [colors.HexColor('#F5F5F5'), colors.white]
        
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E4057')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
            
            # Alternating row colors
            *([('ROWBACKGROUNDS', (0, 1), (-1, -1), row_colors)])
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph("No sales data found for the selected period.", styles["Normal"]))
    
    # Build the PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer
    buffer.seek(0)
    
    # Create the response
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.pdf'
    
    return response

def generate_stock_report(start_date, end_date):
    # Query stock additions
    stock_data = db.session.query(
        StockAddition.id,
        Product.name.label('product_name'),
        Product.sku,
        Product.category,
        StockAddition.quantity_added,
        StockAddition.date_added,
        User.username.label('added_by'),
        StockAddition.old_cost_price,
        StockAddition.new_cost_price,
        StockAddition.old_selling_price,
        StockAddition.new_selling_price,
        StockAddition.price_change_reason
    ).join(Product).join(User).filter(
        and_(StockAddition.date_added >= start_date, StockAddition.date_added <= end_date)
    ).all()
    
    # Current stock levels
    current_stock = Product.query.all()
    
    # Generate PDF
    buffer = BytesIO()
    doc = DocTemplateWithFooter(
        buffer, 
        pagesize=landscape(letter),
        leftMargin=0.75*inch,
        rightMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=1*inch
    )
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2E4057')
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#4A6572')
    )
    
    # Company header
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#546E7A')
    )
    
    # Add company header
    elements.append(Paragraph("CKS BUSINESS MANAGEMENT SYSTEM", header_style))
    elements.append(Spacer(1, 0.1*inch))
    
    # Add title
    elements.append(Paragraph("STOCK MOVEMENT REPORT", title_style))
    elements.append(Paragraph(f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}", subtitle_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Report generation info
    report_info_style = ParagraphStyle(
        'ReportInfo',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_RIGHT,
        textColor=colors.HexColor('#546E7A')
    )
    
    elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", report_info_style))
    elements.append(Paragraph(f"Generated by: {current_user.username}", report_info_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Summary section
    elements.append(Paragraph("REPORT SUMMARY", subtitle_style))
    
    # Calculate summary data
    total_products = len(current_stock)
    total_stock_value = sum([product.quantity_in_stock * product.cost_price for product in current_stock])
    total_stock_units = sum([product.quantity_in_stock for product in current_stock])
    total_additions = len(stock_data)
    total_added_units = sum([stock.quantity_added for stock in stock_data])
    
    # Create summary table
    summary_data = [
        ['Total Products', f'{total_products:,}'],
        ['Total Stock Value', f'UGX {total_stock_value:,.2f}'],
        ['Total Stock Units', f'{total_stock_units:,}'],
        ['Total Stock Additions', f'{total_additions:,}'],
        ['Total Units Added', f'{total_added_units:,}']
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5F5F5')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#424242')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0'))
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Stock additions section
    elements.append(Paragraph("STOCK ADDITIONS", subtitle_style))
    
    if stock_data:
        # Create stock additions table
        headers = ['ID', 'Product', 'SKU', 'Category', 'Qty Added', 'Date', 'Added By', 'Price Changes']
        data = [headers]
        
        for stock in stock_data:
            price_changes = ""
            if stock.old_cost_price != stock.new_cost_price and stock.new_cost_price is not None:
                price_changes += f"Cost: UGX {stock.old_cost_price:.2f} → UGX {stock.new_cost_price:.2f}"
            if stock.old_selling_price != stock.new_selling_price and stock.new_selling_price is not None:
                if price_changes:
                    price_changes += "<br/>"
                price_changes += f"Sell: UGX {stock.old_selling_price:.2f} → UGX {stock.new_selling_price:.2f}"
            
            data.append([
                str(stock.id),
                stock.product_name,
                stock.sku,
                stock.category,
                str(stock.quantity_added),
                stock.date_added.strftime('%Y-%m-%d %H:%M'),
                stock.added_by,
                price_changes
            ])
        
        # Create table with alternating row colors
        table = Table(data, colWidths=[0.5*inch, 1.5*inch, 0.8*inch, 1*inch, 0.8*inch, 1.2*inch, 1*inch, 1.5*inch])
        
        # Style the table
        row_colors = [colors.HexColor('#F5F5F5'), colors.white]
        
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E4057')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
            
            # Alternating row colors
            *([('ROWBACKGROUNDS', (0, 1), (-1, -1), row_colors)])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
    
    # Current stock levels section
    elements.append(Paragraph("CURRENT STOCK LEVELS", subtitle_style))
    
    # Create current stock table
    headers = ['Product', 'SKU', 'Category', 'Current Stock', 'Cost Price', 'Selling Price', 'Total Value', 'Status']
    data = [headers]
    
    for product in current_stock:
        # Determine stock status
        if product.quantity_in_stock == 0:
            status = "Out of Stock"
            status_color = colors.HexColor('#FF5252')
        elif product.quantity_in_stock <= product.low_stock_threshold:
            status = "Low Stock"
            status_color = colors.HexColor('#FFC107')
        else:
            status = "In Stock"
            status_color = colors.HexColor('#4CAF50')
        
        data.append([
            product.name,
            product.sku,
            product.category,
            str(product.quantity_in_stock),
            f'UGX {product.cost_price:,.2f}',
            f'UGX {product.selling_price:,.2f}',
            f'UGX {product.total_value:,.2f}',
            status
        ])
    
    # Create table with alternating row colors
    table = Table(data, colWidths=[1.5*inch, 0.8*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch, 1*inch, 0.8*inch])
    
    # Style the table
    row_colors = [colors.HexColor('#F5F5F5'), colors.white]
    
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E4057')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data styling
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E0E0E0')),
        
        # Alternating row colors
        *([('ROWBACKGROUNDS', (0, 1), (-1, -1), row_colors)])
    ]))
    
    # Apply conditional formatting for stock status
    for i, row in enumerate(data[1:], start=1):
        if row[7] == "Out of Stock":
            table.setStyle(TableStyle([
                ('TEXTCOLOR', (7, i), (7, i), colors.HexColor('#FF5252')),
                ('FONTNAME', (7, i), (7, i), 'Helvetica-Bold')
            ]))
        elif row[7] == "Low Stock":
            table.setStyle(TableStyle([
                ('TEXTCOLOR', (7, i), (7, i), colors.HexColor('#FFC107')),
                ('FONTNAME', (7, i), (7, i), 'Helvetica-Bold')
            ]))
        else:
            table.setStyle(TableStyle([
                ('TEXTCOLOR', (7, i), (7, i), colors.HexColor('#4CAF50')),
                ('FONTNAME', (7, i), (7, i), 'Helvetica-Bold')
            ]))
    
    elements.append(table)
    
    # Build the PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer
    buffer.seek(0)
    
    # Create the response
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=stock_report_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.pdf'
    
    return response

def generate_sales_excel(start_date, end_date):
    # Query sales data
    sales_data = db.session.query(
        Sale.id,
        Product.name.label('product_name'),
        Product.sku,
        Product.category,
        Sale.quantity_sold,
        Sale.price_per_unit,
        Sale.total_amount,
        Sale.timestamp,
        User.username.label('employee')
    ).join(Product).join(User).filter(
        and_(Sale.timestamp >= start_date, Sale.timestamp <= end_date)
    ).all()
    
    # Create a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    # Create a workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
    title_font = Font(name='Calibri', size=16, bold=True, color='2E4057')
    subtitle_font = Font(name='Calibri', size=12, bold=True, color='4A6572')
    data_font = Font(name='Calibri', size=10)
    currency_format = '#,##0.00'
    date_format = 'YYYY-MM-DD HH:MM'
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add company header
    ws.merge_cells('A1:I1')
    ws['A1'] = "CKS BUSINESS MANAGEMENT SYSTEM"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='546E7A')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add title
    ws.merge_cells('A2:I2')
    ws['A2'] = "SALES REPORT"
    ws['A2'].font = title_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Add period
    ws.merge_cells('A3:I3')
    ws['A3'] = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    ws['A3'].font = subtitle_font
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add report generation info
    ws['G4'] = "Generated on:"
    ws['G4'].font = Font(name='Calibri', size=10, color='546E7A')
    ws['H4'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    ws['H4'].font = Font(name='Calibri', size=10)
    
    ws['G5'] = "Generated by:"
    ws['G5'].font = Font(name='Calibri', size=10, color='546E7A')
    ws['H5'] = current_user.username
    ws['H5'].font = Font(name='Calibri', size=10)
    
    # Add summary section
    ws['A7'] = "REPORT SUMMARY"
    ws['A7'].font = subtitle_font
    
    # Calculate summary data
    total_sales = sum([sale.total_amount for sale in sales_data])
    total_quantity = sum([sale.quantity_sold for sale in sales_data])
    total_transactions = len(sales_data)
    
    # Create summary table
    summary_data = [
        ['Total Sales Amount', f'UGX {total_sales:,.2f}'],
        ['Total Units Sold', f'{total_quantity:,} units'],
        ['Number of Transactions', f'{total_transactions:,}'],
        ['Average Transaction Value', f'UGX {total_sales/total_transactions if total_transactions > 0 else 0:,.2f}']
    ]
    
    # Add summary data to worksheet
    for i, (label, value) in enumerate(summary_data, start=8):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = Font(name='Calibri', size=11)
        ws[f'B{i}'].number_format = currency_format if 'UGX' in value else 'General'
    
    # Add sales details section
    start_row = 13
    ws[f'A{start_row}'] = "SALES DETAILS"
    ws[f'A{start_row}'].font = subtitle_font
    
    # Add headers
    headers = ['ID', 'Product', 'SKU', 'Category', 'Qty', 'Price/Unit', 'Total', 'Date', 'Employee']
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row+1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for i, sale in enumerate(sales_data, start=start_row+2):
        ws.cell(row=i, column=1, value=sale.id).font = data_font
        ws.cell(row=i, column=2, value=sale.product_name).font = data_font
        ws.cell(row=i, column=3, value=sale.sku).font = data_font
        ws.cell(row=i, column=4, value=sale.category).font = data_font
        ws.cell(row=i, column=5, value=sale.quantity_sold).font = data_font
        ws.cell(row=i, column=5).number_format = '#,##0'
        ws.cell(row=i, column=6, value=sale.price_per_unit).font = data_font
        ws.cell(row=i, column=6).number_format = currency_format
        ws.cell(row=i, column=7, value=sale.total_amount).font = data_font
        ws.cell(row=i, column=7).number_format = currency_format
        ws.cell(row=i, column=8, value=sale.timestamp).font = data_font
        ws.cell(row=i, column=8).number_format = date_format
        ws.cell(row=i, column=9, value=sale.employee).font = data_font
        
        # Add borders to data cells
        for j in range(1, 10):
            ws.cell(row=i, column=j).border = thin_border
    
    # Adjust column widths
    column_widths = [8, 20, 12, 15, 8, 12, 12, 18, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add footer
    footer_row = ws.max_row + 2
    ws.merge_cells(f'A{footer_row}:I{footer_row}')
    ws[f'A{footer_row}'] = "© Copyright CKS Tech | Contact: +256 755261254 | Email: kephacheps55@gmail.com"
    ws[f'A{footer_row}'].font = Font(name='Calibri', size=9, color='546E7A')
    ws[f'A{footer_row}'].alignment = Alignment(horizontal='center')
    
    # Save the workbook
    wb.save(temp_file.name)
    
    # Read the file and create response
    with open(temp_file.name, 'rb') as f:
        file_data = f.read()
    
    # Delete the temporary file
    os.unlink(temp_file.name)
    
    # Create the response
    response = make_response(file_data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=sales_report_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.xlsx'
    
    return response

def generate_stock_excel(start_date, end_date):
    # Query stock additions
    stock_data = db.session.query(
        StockAddition.id,
        Product.name.label('product_name'),
        Product.sku,
        Product.category,
        StockAddition.quantity_added,
        StockAddition.date_added,
        User.username.label('added_by'),
        StockAddition.old_cost_price,
        StockAddition.new_cost_price,
        StockAddition.old_selling_price,
        StockAddition.new_selling_price,
        StockAddition.price_change_reason
    ).join(Product).join(User).filter(
        and_(StockAddition.date_added >= start_date, StockAddition.date_added <= end_date)
    ).all()
    
    # Current stock levels
    current_stock = Product.query.all()
    
    # Create a temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_file.close()
    
    # Create a workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Movement Report"
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
    title_font = Font(name='Calibri', size=16, bold=True, color='2E4057')
    subtitle_font = Font(name='Calibri', size=12, bold=True, color='4A6572')
    data_font = Font(name='Calibri', size=10)
    currency_format = '#,##0.00'
    date_format = 'YYYY-MM-DD HH:MM'
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add company header
    ws.merge_cells('A1:L1')
    ws['A1'] = "CKS BUSINESS MANAGEMENT SYSTEM"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='546E7A')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add title
    ws.merge_cells('A2:L2')
    ws['A2'] = "STOCK MOVEMENT REPORT"
    ws['A2'].font = title_font
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Add period
    ws.merge_cells('A3:L3')
    ws['A3'] = f"Period: {start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}"
    ws['A3'].font = subtitle_font
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Add report generation info
    ws['J4'] = "Generated on:"
    ws['J4'].font = Font(name='Calibri', size=10, color='546E7A')
    ws['K4'] = datetime.now().strftime('%B %d, %Y at %I:%M %p')
    ws['K4'].font = Font(name='Calibri', size=10)
    
    ws['J5'] = "Generated by:"
    ws['J5'].font = Font(name='Calibri', size=10, color='546E7A')
    ws['K5'] = current_user.username
    ws['K5'].font = Font(name='Calibri', size=10)
    
    # Add summary section
    ws['A7'] = "REPORT SUMMARY"
    ws['A7'].font = subtitle_font
    
    # Calculate summary data
    total_products = len(current_stock)
    total_stock_value = sum([product.quantity_in_stock * product.cost_price for product in current_stock])
    total_stock_units = sum([product.quantity_in_stock for product in current_stock])
    total_additions = len(stock_data)
    total_added_units = sum([stock.quantity_added for stock in stock_data])
    
    # Create summary table
    summary_data = [
        ['Total Products', f'{total_products:,}'],
        ['Total Stock Value', f'UGX {total_stock_value:,.2f}'],
        ['Total Stock Units', f'{total_stock_units:,}'],
        ['Total Stock Additions', f'{total_additions:,}'],
        ['Total Units Added', f'{total_added_units:,}']
    ]
    
    # Add summary data to worksheet
    for i, (label, value) in enumerate(summary_data, start=8):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'B{i}'] = value
        ws[f'B{i}'].font = Font(name='Calibri', size=11)
        ws[f'B{i}'].number_format = currency_format if 'UGX' in value else 'General'
    
    # Add stock additions section
    start_row = 14
    ws[f'A{start_row}'] = "STOCK ADDITIONS"
    ws[f'A{start_row}'].font = subtitle_font
    
    # Add headers
    headers = ['ID', 'Product', 'SKU', 'Category', 'Qty Added', 'Date', 'Added By', 'Old Cost', 'New Cost', 'Old Sell', 'New Sell', 'Reason']
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row+1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for i, stock in enumerate(stock_data, start=start_row+2):
        ws.cell(row=i, column=1, value=stock.id).font = data_font
        ws.cell(row=i, column=2, value=stock.product_name).font = data_font
        ws.cell(row=i, column=3, value=stock.sku).font = data_font
        ws.cell(row=i, column=4, value=stock.category).font = data_font
        ws.cell(row=i, column=5, value=stock.quantity_added).font = data_font
        ws.cell(row=i, column=5).number_format = '#,##0'
        ws.cell(row=i, column=6, value=stock.date_added).font = data_font
        ws.cell(row=i, column=6).number_format = date_format
        ws.cell(row=i, column=7, value=stock.added_by).font = data_font
        ws.cell(row=i, column=8, value=stock.old_cost_price).font = data_font
        ws.cell(row=i, column=8).number_format = currency_format
        ws.cell(row=i, column=9, value=stock.new_cost_price).font = data_font
        ws.cell(row=i, column=9).number_format = currency_format
        ws.cell(row=i, column=10, value=stock.old_selling_price).font = data_font
        ws.cell(row=i, column=10).number_format = currency_format
        ws.cell(row=i, column=11, value=stock.new_selling_price).font = data_font
        ws.cell(row=i, column=11).number_format = currency_format
        ws.cell(row=i, column=12, value=stock.price_change_reason).font = data_font
        
        # Add borders to data cells
        for j in range(1, 13):
            ws.cell(row=i, column=j).border = thin_border
    
    # Add current stock levels section
    start_row = ws.max_row + 3
    ws[f'A{start_row}'] = "CURRENT STOCK LEVELS"
    ws[f'A{start_row}'].font = subtitle_font
    
    # Add headers
    headers = ['Product', 'SKU', 'Category', 'Current Stock', 'Cost Price', 'Selling Price', 'Total Value', 'Status']
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row+1, column=i, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for i, product in enumerate(current_stock, start=start_row+2):
        # Determine stock status
        if product.quantity_in_stock == 0:
            status = "Out of Stock"
            status_color = 'FFFF5252'  # Red
        elif product.quantity_in_stock <= product.low_stock_threshold:
            status = "Low Stock"
            status_color = 'FFFFC107'  # Yellow
        else:
            status = "In Stock"
            status_color = 'FF4CAF50'  # Green
        
        ws.cell(row=i, column=1, value=product.name).font = data_font
        ws.cell(row=i, column=2, value=product.sku).font = data_font
        ws.cell(row=i, column=3, value=product.category).font = data_font
        ws.cell(row=i, column=4, value=product.quantity_in_stock).font = data_font
        ws.cell(row=i, column=4).number_format = '#,##0'
        ws.cell(row=i, column=5, value=product.cost_price).font = data_font
        ws.cell(row=i, column=5).number_format = currency_format
        ws.cell(row=i, column=6, value=product.selling_price).font = data_font
        ws.cell(row=i, column=6).number_format = currency_format
        ws.cell(row=i, column=7, value=product.total_value).font = data_font
        ws.cell(row=i, column=7).number_format = currency_format
        ws.cell(row=i, column=8, value=status).font = Font(name='Calibri', size=10, bold=True)
        ws.cell(row=i, column=8).fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
        
        # Add borders to data cells
        for j in range(1, 9):
            ws.cell(row=i, column=j).border = thin_border
    
    # Adjust column widths
    column_widths = [20, 12, 15, 12, 12, 12, 12, 15, 12, 12, 12, 20]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add footer
    footer_row = ws.max_row + 2
    ws.merge_cells(f'A{footer_row}:L{footer_row}')
    ws[f'A{footer_row}'] = "© Copyright CKS Tech | Contact: +256 755261254 | Email: kephacheps55@gmail.com"
    ws[f'A{footer_row}'].font = Font(name='Calibri', size=9, color='546E7A')
    ws[f'A{footer_row}'].alignment = Alignment(horizontal='center')
    
    # Save the workbook
    wb.save(temp_file.name)
    
    # Read the file and create response
    with open(temp_file.name, 'rb') as f:
        file_data = f.read()
    
    # Delete the temporary file
    os.unlink(temp_file.name)
    
    # Create the response
    response = make_response(file_data)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=stock_report_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.xlsx'
    
    return response